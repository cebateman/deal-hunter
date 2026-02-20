"""
Deal Hunter — BizBuySell Scraper & Scoring Engine
Scrapes listings, scores against acquisition criteria, outputs to Excel + email digest.
"""

import json
import re
import time
import subprocess
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional
from urllib.parse import quote

# --- ACQUISITION CRITERIA ---
CRITERIA = {
    "ev_min": 1_000_000,
    "ev_max": 5_000_000,
    "revenue_min": 2_000_000,
    "revenue_max": 15_000_000,
    "ebitda_min": 300_000,
    "max_multiple": 4.0,
    "geography": "United States",
    "preferred_traits": [
        "recurring_revenue", "regulatory_moat", "labor_accessible",
        "high_switching_costs", "non_cyclical", "unglamorous", "essential_service"
    ],
    "avoid_traits": [
        "commodity_exposure", "cyclical_demand", "specialized_labor_required",
        "asset_light_digital", "construction_tied"
    ],
    "target_industries": [
        "Water Treatment", "Fire Protection", "Elevator Maintenance",
        "Environmental Remediation", "Commercial Laundry", "Meat Processing",
        "Produce Packing", "Fresh-Cut Vegetables", "Hide/Leather Tanning",
        "Pallet Recycling", "Textile Recycling", "Seafood Processing",
        "Contract Packaging", "Industrial Parts Cleaning", "Janitorial Services",
        "Industrial Refrigeration", "Demolition & Salvage"
    ],
    # BizBuySell search categories that map to your interests
    "bizbuysell_categories": [
        "service-businesses",
        "manufacturing-businesses",
        "wholesale-and-distributor-businesses",
    ],
    # Keywords to search within each category
    "search_keywords": [
        "laundry", "fire sprinkler", "fire protection", "elevator",
        "remediation", "abatement", "water treatment", "meat processing",
        "produce", "fresh cut", "seafood", "fish processing",
        "pallet", "textile", "recycling", "packaging", "co-packing",
        "industrial cleaning", "parts cleaning", "degreasing",
        "janitorial", "commercial cleaning", "refrigeration",
        "tanning", "hide", "leather processing",
        "demolition", "environmental services",
    ],
}

# Keywords that signal traits (used in description analysis)
TRAIT_KEYWORDS = {
    "recurring_revenue": ["contract", "recurring", "subscription", "auto-renew", "monthly", "annual contract", "repeat", "retainer"],
    "regulatory_moat": ["licensed", "permit", "certified", "EPA", "FDA", "USDA", "OSHA", "regulated", "compliance", "inspection", "certification"],
    "labor_accessible": ["train", "no experience", "entry level", "on-the-job", "manual", "production line", "floor worker", "trainable", "unskilled"],
    "high_switching_costs": ["switching cost", "long-term contract", "auto-renew", "embedded", "sole provider", "exclusive"],
    "non_cyclical": ["essential", "recession", "steady", "consistent", "stable demand", "non-discretionary", "maintenance", "required by law", "mandatory"],
    "unglamorous": ["niche", "overlooked", "few competitors", "no one wants", "unglamorous"],
    "essential_service": ["essential", "critical", "life safety", "health", "food", "water", "maintenance", "compliance", "required"],
    "commodity_exposure": ["commodity", "spot price", "market price", "lumber", "steel price", "oil price"],
    "cyclical_demand": ["cyclical", "seasonal", "construction cycle", "housing market", "real estate dependent"],
    "specialized_labor_required": ["engineer required", "degree required", "specialized certification", "hard to hire"],
    "construction_tied": ["construction", "new build", "housing", "real estate development"],
}

# Industry classification keywords
INDUSTRY_KEYWORDS = {
    "Commercial Laundry": ["laundry", "linen", "uniform service", "textile cleaning"],
    "Fire Protection": ["fire sprinkler", "fire protection", "fire suppression", "fire alarm"],
    "Elevator Maintenance": ["elevator", "escalator", "lift maintenance"],
    "Environmental Remediation": ["remediation", "abatement", "asbestos", "mold removal", "lead abatement", "environmental clean"],
    "Water Treatment": ["water treatment", "water purification", "water service"],
    "Meat Processing": ["meat processing", "butcher", "slaughter", "meat packing"],
    "Produce Packing": ["produce", "fresh cut", "vegetable processing", "fruit packing"],
    "Seafood Processing": ["seafood", "fish processing", "fish packing", "shrimp"],
    "Pallet Recycling": ["pallet", "pallet recycl", "pallet repair"],
    "Textile Recycling": ["textile recycl", "rag processing", "fiber recycl"],
    "Contract Packaging": ["co-pack", "contract pack", "packaging service"],
    "Industrial Parts Cleaning": ["parts cleaning", "degreasing", "industrial cleaning"],
    "Janitorial Services": ["janitorial", "commercial cleaning", "building maintenance", "custodial"],
    "Industrial Refrigeration": ["refrigeration", "cold storage", "HVAC service", "cooling"],
    "Hide/Leather Tanning": ["tanning", "hide", "leather processing", "fur dressing"],
    "Demolition & Salvage": ["demolition", "salvage", "deconstruction"],
}


@dataclass
class Deal:
    title: str = ""
    url: str = ""
    location: str = ""
    asking_price: Optional[float] = None
    revenue: Optional[float] = None
    ebitda: Optional[float] = None
    cash_flow_sde: Optional[float] = None
    year_established: Optional[int] = None
    employees: Optional[int] = None
    description: str = ""
    source: str = "BizBuySell"
    industry: str = "Unknown"
    date_found: str = ""
    traits: list = field(default_factory=list)
    avoid_traits: list = field(default_factory=list)
    score: int = 0
    multiple: Optional[float] = None
    broker: str = ""
    listing_id: str = ""
    category: str = ""


def parse_money(text: str) -> Optional[float]:
    """Parse money string like '$1,234,567' or '$1.2M' into float."""
    if not text or text.strip().lower() in ("not disclosed", "n/a", ""):
        return None
    text = text.strip().replace(",", "").replace("$", "")
    try:
        if "m" in text.lower():
            return float(text.lower().replace("m", "")) * 1_000_000
        elif "k" in text.lower():
            return float(text.lower().replace("k", "")) * 1_000
        return float(text)
    except (ValueError, TypeError):
        return None


def classify_industry(title: str, description: str) -> str:
    """Classify a deal into one of our target industries based on keywords."""
    text = f"{title} {description}".lower()
    for industry, keywords in INDUSTRY_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in text:
                return industry
    return "Other"


def detect_traits(description: str, title: str) -> tuple[list, list]:
    """Analyze description to detect positive and negative traits."""
    text = f"{title} {description}".lower()
    positive = []
    negative = []

    for trait, keywords in TRAIT_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in text:
                if trait in CRITERIA["preferred_traits"]:
                    if trait not in positive:
                        positive.append(trait)
                elif trait in CRITERIA["avoid_traits"]:
                    if trait not in negative:
                        negative.append(trait)
                break

    return positive, negative


def score_deal(deal: Deal) -> int:
    """Score a deal 0-100 based on criteria match."""
    score = 0
    max_possible = 100

    # Trait scoring (50% weight)
    trait_score = 0
    max_trait = len(CRITERIA["preferred_traits"]) * 10
    for t in deal.traits:
        if t in CRITERIA["preferred_traits"]:
            trait_score += 10
    for t in deal.avoid_traits:
        if t in CRITERIA["avoid_traits"]:
            trait_score -= 15
    trait_score = max(0, min(100, (trait_score / max_trait) * 100 if max_trait > 0 else 0))

    # Multiple scoring (30% weight)
    multiple_score = 0
    if deal.multiple is not None:
        if deal.multiple <= 2.5:
            multiple_score = 100
        elif deal.multiple <= 3.0:
            multiple_score = 90
        elif deal.multiple <= 3.5:
            multiple_score = 75
        elif deal.multiple <= 4.0:
            multiple_score = 50
        else:
            multiple_score = 0

    # Industry match (20% weight)
    industry_score = 100 if deal.industry in CRITERIA["target_industries"] else 20

    total = trait_score * 0.5 + multiple_score * 0.3 + industry_score * 0.2
    return min(100, round(total))


def compute_multiple(deal: Deal) -> Optional[float]:
    """Compute asking price / EBITDA multiple."""
    earnings = deal.ebitda or deal.cash_flow_sde
    if deal.asking_price and earnings and earnings > 0:
        return round(deal.asking_price / earnings, 2)
    return None


def passes_financial_filters(deal: Deal) -> bool:
    """Check if deal meets basic financial criteria."""
    # Must have asking price in range
    if deal.asking_price:
        if deal.asking_price < CRITERIA["ev_min"] or deal.asking_price > CRITERIA["ev_max"]:
            return False

    # Check EBITDA minimum
    earnings = deal.ebitda or deal.cash_flow_sde
    if earnings and earnings < CRITERIA["ebitda_min"]:
        return False

    # Check multiple cap
    if deal.multiple and deal.multiple > CRITERIA["max_multiple"]:
        return False

    return True


def generate_bizbuysell_urls() -> list[str]:
    """Generate BizBuySell search URLs for all relevant categories with price filters."""
    urls = []
    base = "https://www.bizbuysell.com"

    for category in CRITERIA["bizbuysell_categories"]:
        # Base category URL with price range $1M-$5M
        # BizBuySell uses query params for filters
        url = f"{base}/{category}-for-sale/"
        urls.append(url)

    # Also generate keyword-specific searches
    for keyword in CRITERIA["search_keywords"]:
        encoded = quote(keyword)
        url = f"{base}/businesses-for-sale/?q={encoded}"
        urls.append(url)

    return urls


def build_search_urls_with_filters() -> list[str]:
    """Build BizBuySell URLs with price and financial filters baked in."""
    urls = []
    base = "https://www.bizbuysell.com"

    # Category-based searches with price filters
    for cat in CRITERIA["bizbuysell_categories"]:
        # Price range $1M to $5M
        urls.append(
            f"{base}/{cat}-for-sale/"
            f"?q=bGM9SmtjOU1UQW1RejFWVXc9PQ%3D%3D"  # US location
            f"&price_min=1000000&price_max=5000000"
        )

    # Keyword searches
    for kw in CRITERIA["search_keywords"][:15]:  # Limit to avoid rate limiting
        urls.append(
            f"{base}/businesses-for-sale/"
            f"?kw={quote(kw)}"
            f"&price_min=1000000&price_max=5000000"
        )

    return urls


# --- HTML PARSING (works with requests + BeautifulSoup, no browser needed) ---

def parse_listing_card(card_html: str, source_url: str = "") -> Optional[Deal]:
    """Parse a BizBuySell listing card HTML into a Deal object."""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(card_html, "html.parser")

    deal = Deal()
    deal.date_found = datetime.now().strftime("%Y-%m-%d")
    deal.source = "BizBuySell"

    # Title and URL
    title_el = soup.select_one("a.diamond-header, h3 a, .listing-title a, a[href*='business-opportunity']")
    if title_el:
        deal.title = title_el.get_text(strip=True)
        href = title_el.get("href", "")
        deal.url = f"https://www.bizbuysell.com{href}" if href.startswith("/") else href
        # Extract listing ID from URL
        match = re.search(r"/(\d+)/?$", deal.url)
        if match:
            deal.listing_id = match.group(1)

    # Location
    loc_el = soup.select_one(".listing-location, .location")
    if loc_el:
        deal.location = loc_el.get_text(strip=True)

    # Financials
    text = soup.get_text()

    # Asking price
    price_match = re.search(r"\$[\d,]+(?:\.\d+)?", text)
    if price_match:
        deal.asking_price = parse_money(price_match.group())

    # EBITDA
    ebitda_match = re.search(r"EBITDA:\s*\$?([\d,]+)", text)
    if ebitda_match:
        deal.ebitda = parse_money(ebitda_match.group(1))

    # Cash Flow
    cf_match = re.search(r"Cash Flow[^:]*:\s*\$?([\d,]+)", text)
    if cf_match:
        deal.cash_flow_sde = parse_money(cf_match.group(1))

    # Revenue
    rev_match = re.search(r"(?:Revenue|Gross Revenue)[^:]*:\s*\$?([\d,]+)", text)
    if rev_match:
        deal.revenue = parse_money(rev_match.group(1))

    # Description
    desc_el = soup.select_one(".listing-description, .listing-text, p")
    if desc_el:
        deal.description = desc_el.get_text(strip=True)[:500]

    return deal


def process_deal(deal: Deal) -> Deal:
    """Enrich a deal with classification, traits, score."""
    deal.industry = classify_industry(deal.title, deal.description)
    deal.traits, deal.avoid_traits = detect_traits(deal.description, deal.title)
    deal.multiple = compute_multiple(deal)
    deal.score = score_deal(deal)
    return deal


# --- PLAYWRIGHT SCRAPER ---

SCRAPER_SCRIPT = '''
const {{ chromium }} = require('playwright');

(async () => {{
    const browser = await chromium.launch({{ headless: true }});
    const page = await browser.newPage();

    const urls = {urls_json};
    const allListings = [];
    const seen = new Set();

    for (const url of urls) {{
        try {{
            await page.goto(url, {{ timeout: 30000 }});
            await page.waitForTimeout(2000);

            const listings = await page.$$eval('.listing, .search-result, [class*="listing"]', cards => {{
                return cards.map(card => {{
                    const titleEl = card.querySelector('a[href*="business-opportunity"]') || card.querySelector('h3 a') || card.querySelector('.diamond-header');
                    const title = titleEl?.textContent?.trim() || '';
                    const href = titleEl?.getAttribute('href') || '';
                    const text = card.textContent || '';
                    return {{ title, href, text, html: card.outerHTML }};
                }});
            }});

            for (const l of listings) {{
                if (l.title && !seen.has(l.title)) {{
                    seen.add(l.title);
                    allListings.push(l);
                }}
            }}
        }} catch (e) {{
            console.error(`Error on ${{url}}: ${{e.message}}`);
        }}
    }}

    console.log(JSON.stringify(allListings));
    await browser.close();
}})();
'''


# --- EXCEL OUTPUT ---

def write_deals_to_excel(deals: list[Deal], filename: str = "deal_hunter_tracker.xlsx"):
    """Write scored deals to a formatted Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- DEALS SHEET ---
    ws = wb.active
    ws.title = "Deal Pipeline"

    headers = [
        "Score", "Title", "Industry", "Location", "Asking Price",
        "Revenue", "EBITDA/SDE", "Multiple", "Year Est.", "Employees",
        "Positive Traits", "Red Flags", "Description", "Source URL",
        "Date Found", "Rating", "Notes"
    ]

    # Header styling
    header_fill = PatternFill("solid", fgColor="1a1f2e")
    header_font = Font(name="Arial", bold=True, color="f59e0b", size=11)
    border = Border(bottom=Side(style="thin", color="333333"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Column widths
    widths = [8, 40, 22, 20, 14, 14, 14, 10, 10, 10, 30, 20, 60, 45, 12, 15, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"].width = w

    # Data rows
    green_font = Font(name="Arial", color="059669", size=10)
    red_font = Font(name="Arial", color="ef4444", size=10)
    normal_font = Font(name="Arial", size=10)
    money_fmt = "$#,##0"
    multiple_fmt = "0.00x"

    sorted_deals = sorted(deals, key=lambda d: d.score, reverse=True)

    for row_idx, deal in enumerate(sorted_deals, 2):
        # Score with conditional color
        score_cell = ws.cell(row=row_idx, column=1, value=deal.score)
        if deal.score >= 80:
            score_cell.font = Font(name="Arial", bold=True, color="059669", size=11)
        elif deal.score >= 60:
            score_cell.font = Font(name="Arial", bold=True, color="2563eb", size=11)
        elif deal.score >= 40:
            score_cell.font = Font(name="Arial", bold=True, color="d97706", size=11)
        else:
            score_cell.font = Font(name="Arial", bold=True, color="6b7280", size=11)
        score_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=2, value=deal.title).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row_idx, column=3, value=deal.industry).font = normal_font
        ws.cell(row=row_idx, column=4, value=deal.location).font = normal_font

        price_cell = ws.cell(row=row_idx, column=5, value=deal.asking_price)
        price_cell.number_format = money_fmt
        price_cell.font = normal_font

        rev_cell = ws.cell(row=row_idx, column=6, value=deal.revenue)
        rev_cell.number_format = money_fmt
        rev_cell.font = normal_font

        earnings = deal.ebitda or deal.cash_flow_sde
        earn_cell = ws.cell(row=row_idx, column=7, value=earnings)
        earn_cell.number_format = money_fmt
        earn_cell.font = normal_font

        mult_cell = ws.cell(row=row_idx, column=8, value=deal.multiple)
        if deal.multiple:
            mult_cell.number_format = "0.0\"x\""
            if deal.multiple <= 3.0:
                mult_cell.font = Font(name="Arial", color="059669", bold=True, size=10)
            elif deal.multiple <= 4.0:
                mult_cell.font = Font(name="Arial", color="d97706", size=10)
            else:
                mult_cell.font = Font(name="Arial", color="ef4444", size=10)
        mult_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=9, value=deal.year_established).font = normal_font
        ws.cell(row=row_idx, column=10, value=deal.employees).font = normal_font

        trait_labels = {
            "recurring_revenue": "Recurring Rev", "regulatory_moat": "Reg. Moat",
            "labor_accessible": "Trainable Labor", "high_switching_costs": "High Switch Cost",
            "non_cyclical": "Non-Cyclical", "unglamorous": "Unglamorous",
            "essential_service": "Essential Svc",
        }
        avoid_labels = {
            "commodity_exposure": "Commodity", "cyclical_demand": "Cyclical",
            "specialized_labor_required": "Specialized Labor",
            "asset_light_digital": "Digital", "construction_tied": "Construction",
        }

        traits_str = ", ".join(trait_labels.get(t, t) for t in deal.traits)
        avoid_str = ", ".join(avoid_labels.get(t, t) for t in deal.avoid_traits)

        ws.cell(row=row_idx, column=11, value=traits_str).font = green_font
        ws.cell(row=row_idx, column=12, value=avoid_str).font = red_font
        ws.cell(row=row_idx, column=13, value=deal.description[:300]).font = Font(name="Arial", size=9, color="666666")
        ws.cell(row=row_idx, column=14, value=deal.url).font = Font(name="Arial", size=9, color="2563eb")
        ws.cell(row=row_idx, column=15, value=deal.date_found).font = normal_font
        ws.cell(row=row_idx, column=16, value="").font = normal_font  # Rating (user fills)
        ws.cell(row=row_idx, column=17, value="").font = normal_font  # Notes (user fills)

        # Alternate row shading
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="f8f9fa")

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- CRITERIA SHEET ---
    cs = wb.create_sheet("Acquisition Criteria")
    cs.column_dimensions["A"].width = 25
    cs.column_dimensions["B"].width = 40

    criteria_data = [
        ("ACQUISITION CRITERIA", ""),
        ("", ""),
        ("Enterprise Value", "$1M – $5M"),
        ("Revenue Range", "$2M – $15M"),
        ("Minimum EBITDA", "$300K"),
        ("Maximum Multiple", "4.0x EBITDA"),
        ("Geography", "Anywhere in US"),
        ("Structure", "Holding Co — Retain/Install Mgmt"),
        ("", ""),
        ("PREFERRED TRAITS", ""),
        ("", "Recurring Revenue"),
        ("", "Regulatory Moat"),
        ("", "Trainable Labor (accessible to immigrants)"),
        ("", "High Switching Costs"),
        ("", "Non-Cyclical Demand"),
        ("", "Unglamorous / Overlooked"),
        ("", "Essential Service"),
        ("", ""),
        ("AVOID", ""),
        ("", "Commodity Price Exposure"),
        ("", "Cyclical Demand"),
        ("", "Specialized Labor Required"),
        ("", "Asset-Light / Digital"),
        ("", "Construction-Tied"),
        ("", ""),
        ("TARGET INDUSTRIES", ""),
    ]
    for ind in CRITERIA["target_industries"]:
        criteria_data.append(("", ind))

    for row_idx, (a, b) in enumerate(criteria_data, 1):
        cell_a = cs.cell(row=row_idx, column=1, value=a)
        cell_b = cs.cell(row=row_idx, column=2, value=b)
        if a and a.isupper():
            cell_a.font = Font(name="Arial", bold=True, size=12, color="f59e0b")
        else:
            cell_a.font = Font(name="Arial", bold=True, size=10)
        cell_b.font = Font(name="Arial", size=10)

    wb.save(filename)
    return filename


# --- EMAIL DIGEST ---

def generate_email_digest(deals: list[Deal], week_date: str = None) -> str:
    """Generate HTML email digest of top deals."""
    if not week_date:
        week_date = datetime.now().strftime("%B %d, %Y")

    sorted_deals = sorted(deals, key=lambda d: d.score, reverse=True)[:25]

    def fmt_money(n):
        if n is None:
            return "N/A"
        if n >= 1_000_000:
            return f"${n/1_000_000:.1f}M"
        return f"${n/1_000:.0f}K"

    deal_rows = ""
    for i, d in enumerate(sorted_deals):
        earnings = d.ebitda or d.cash_flow_sde
        mult_str = f"{d.multiple:.1f}x" if d.multiple else "N/A"
        traits_str = ", ".join(d.traits[:3]) if d.traits else "—"

        bg = "#f8f9fa" if i % 2 == 0 else "#ffffff"
        score_color = "#059669" if d.score >= 80 else "#2563eb" if d.score >= 60 else "#d97706" if d.score >= 40 else "#6b7280"

        deal_rows += f"""
        <tr style="background:{bg}">
            <td style="padding:12px 8px;text-align:center;font-weight:bold;color:{score_color};font-size:16px">{d.score}</td>
            <td style="padding:12px 8px">
                <a href="{d.url}" style="color:#1a1a1a;font-weight:600;text-decoration:none">{d.title}</a><br>
                <span style="color:#6b7280;font-size:12px">{d.industry} · {d.location}</span>
            </td>
            <td style="padding:12px 8px;text-align:right;font-family:monospace">{fmt_money(d.asking_price)}</td>
            <td style="padding:12px 8px;text-align:right;font-family:monospace">{fmt_money(earnings)}</td>
            <td style="padding:12px 8px;text-align:center;font-family:monospace;color:{'#059669' if d.multiple and d.multiple <= 3.5 else '#d97706' if d.multiple and d.multiple <= 4.0 else '#6b7280'}">{mult_str}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f0f0f0;font-family:Arial,Helvetica,sans-serif">
<div style="max-width:700px;margin:20px auto;background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1)">

    <!-- Header -->
    <div style="background:#0c0f14;padding:24px 32px">
        <h1 style="margin:0;color:#f59e0b;font-size:24px">⬡ DEAL HUNTER</h1>
        <p style="margin:4px 0 0;color:#6b7280;font-size:13px;letter-spacing:1px">WEEKLY ACQUISITION DIGEST — {week_date.upper()}</p>
    </div>

    <!-- Summary -->
    <div style="padding:20px 32px;background:#f8f9fa;border-bottom:1px solid #e5e7eb">
        <p style="margin:0;font-size:15px;color:#374151">
            Found <strong>{len(sorted_deals)} deals</strong> matching your criteria this week.
            Top score: <strong style="color:#059669">{sorted_deals[0].score if sorted_deals else 0}</strong> ·
            Avg multiple: <strong>{sum(d.multiple for d in sorted_deals if d.multiple) / max(1, len([d for d in sorted_deals if d.multiple])):.1f}x</strong>
        </p>
    </div>

    <!-- Deal Table -->
    <div style="padding:16px 32px">
        <table style="width:100%;border-collapse:collapse;font-size:14px">
            <thead>
                <tr style="border-bottom:2px solid #e5e7eb">
                    <th style="padding:8px;text-align:center;color:#6b7280;font-size:11px;letter-spacing:1px">SCORE</th>
                    <th style="padding:8px;text-align:left;color:#6b7280;font-size:11px;letter-spacing:1px">DEAL</th>
                    <th style="padding:8px;text-align:right;color:#6b7280;font-size:11px;letter-spacing:1px">ASK</th>
                    <th style="padding:8px;text-align:right;color:#6b7280;font-size:11px;letter-spacing:1px">EBITDA</th>
                    <th style="padding:8px;text-align:center;color:#6b7280;font-size:11px;letter-spacing:1px">MULT</th>
                </tr>
            </thead>
            <tbody>
                {deal_rows}
            </tbody>
        </table>
    </div>

    <!-- Footer -->
    <div style="padding:20px 32px;background:#f8f9fa;border-top:1px solid #e5e7eb">
        <p style="margin:0;font-size:13px;color:#6b7280">
            <strong>Reply to this email with feedback:</strong> Rate deals as Pass / Maybe / Interested / Strong Interest.
            Your feedback sharpens future results.
        </p>
        <p style="margin:8px 0 0;font-size:12px;color:#9ca3af">
            Deal Hunter v1.0 · Sources: BizBuySell · Criteria: Essential services, regulatory moats, trainable labor, ≤4x EBITDA
        </p>
    </div>

</div>
</body>
</html>"""

    return html


def generate_intro_email() -> str:
    """Generate the first introductory email explaining the system."""
    html = """<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f0f0f0;font-family:Arial,Helvetica,sans-serif">
<div style="max-width:700px;margin:20px auto;background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1)">

    <!-- Header -->
    <div style="background:#0c0f14;padding:32px">
        <h1 style="margin:0;color:#f59e0b;font-size:28px">⬡ DEAL HUNTER</h1>
        <p style="margin:8px 0 0;color:#e2e4e9;font-size:16px">Your weekly acquisition opportunity digest</p>
    </div>

    <!-- Body -->
    <div style="padding:32px">
        <p style="font-size:15px;line-height:1.7;color:#374151;margin:0 0 16px">
            Welcome to Deal Hunter — your automated system for finding acquisition opportunities that match your investment thesis.
        </p>

        <h2 style="font-size:17px;color:#1a1a1a;margin:24px 0 12px">What you'll get each week</h2>
        <p style="font-size:15px;line-height:1.7;color:#374151;margin:0 0 16px">
            Every Monday morning, you'll receive an email like this one with <strong>10-25 scored deals</strong> from BizBuySell and other listing platforms. Each deal is automatically scored against your acquisition criteria on a 0-100 scale.
        </p>

        <h2 style="font-size:17px;color:#1a1a1a;margin:24px 0 12px">Your criteria</h2>
        <div style="background:#f8f9fa;border-radius:8px;padding:20px;margin:0 0 16px">
            <table style="width:100%;font-size:14px;border-collapse:collapse">
                <tr><td style="padding:6px 0;color:#6b7280;width:40%">Enterprise Value</td><td style="padding:6px 0;font-weight:600">$1M – $5M</td></tr>
                <tr><td style="padding:6px 0;color:#6b7280">Revenue</td><td style="padding:6px 0;font-weight:600">$2M – $15M</td></tr>
                <tr><td style="padding:6px 0;color:#6b7280">Min EBITDA</td><td style="padding:6px 0;font-weight:600">$300K</td></tr>
                <tr><td style="padding:6px 0;color:#6b7280">Max Multiple</td><td style="padding:6px 0;font-weight:600">4.0x EBITDA</td></tr>
                <tr><td style="padding:6px 0;color:#6b7280">Geography</td><td style="padding:6px 0;font-weight:600">Anywhere in US</td></tr>
            </table>
        </div>

        <h2 style="font-size:17px;color:#1a1a1a;margin:24px 0 12px">What you're looking for</h2>
        <p style="font-size:15px;line-height:1.7;color:#374151;margin:0 0 8px">
            Essential-service businesses with regulatory moats, recurring revenue, and non-cyclical demand. Labor-intensive operations where floor workers can be trained on the job with minimal prior experience — creating employment pathways while building a defensible holding company of overlooked, high-quality cash flow businesses.
        </p>

        <p style="font-size:14px;line-height:1.7;color:#6b7280;margin:0 0 16px">
            <strong>Target industries:</strong> Commercial laundry, fire protection, elevator maintenance, environmental remediation, meat/seafood/produce processing, pallet & textile recycling, contract packaging, industrial parts cleaning, water treatment, and similar.
        </p>

        <h2 style="font-size:17px;color:#1a1a1a;margin:24px 0 12px">How to give feedback</h2>
        <p style="font-size:15px;line-height:1.7;color:#374151;margin:0 0 16px">
            Just <strong>reply to any weekly email</strong> with your ratings. Use these labels:
        </p>
        <div style="background:#f8f9fa;border-radius:8px;padding:16px 20px;margin:0 0 16px">
            <p style="margin:4px 0;font-size:14px"><span style="color:#059669;font-weight:bold">🔥 Strong Interest</span> — I want to pursue this</p>
            <p style="margin:4px 0;font-size:14px"><span style="color:#2563eb;font-weight:bold">👍 Interested</span> — Worth a closer look</p>
            <p style="margin:4px 0;font-size:14px"><span style="color:#d97706;font-weight:bold">🤔 Maybe</span> — Not sure yet</p>
            <p style="margin:4px 0;font-size:14px"><span style="color:#6b7280;font-weight:bold">👎 Pass</span> — Not for me (tell me why!)</p>
        </div>
        <p style="font-size:15px;line-height:1.7;color:#374151;margin:0 0 16px">
            Your feedback sharpens the scoring algorithm over time. The more you rate, the better the deals get. The "why" on passes is especially valuable — it helps eliminate entire categories you don't want.
        </p>

        <h2 style="font-size:17px;color:#1a1a1a;margin:24px 0 12px">Your deal tracker</h2>
        <p style="font-size:15px;line-height:1.7;color:#374151;margin:0">
            All deals and your feedback are saved in a Google Sheet that's attached to each email. This is your persistent database — you can add notes, change ratings, and track deals across weeks.
        </p>
    </div>

    <!-- Footer -->
    <div style="padding:20px 32px;background:#0c0f14">
        <p style="margin:0;font-size:13px;color:#6b7280">
            Deal Hunter v1.0 · Built for Christian Ellis-Bateman · christianellisbateman@gmail.com
        </p>
    </div>

</div>
</body>
</html>"""

    return html


# --- MAIN ---

if __name__ == "__main__":
    # For demo/testing: create sample deals and generate outputs
    sample_deals = [
        Deal(
            title="ABC Commercial Laundry Services",
            location="Memphis, TN",
            asking_price=2_200_000,
            revenue=3_800_000,
            ebitda=620_000,
            description="Full-service commercial laundry serving 40+ hotel and restaurant accounts. Industrial washers/dryers, established routes. Owner retiring. 12 floor workers trained on-site. Contracts average 3+ years.",
            url="https://www.bizbuysell.com/business-opportunity/example/1234",
            year_established=2008,
            employees=15,
        ),
        Deal(
            title="Southeastern Fire Sprinkler Co.",
            location="Atlanta, GA",
            asking_price=3_100_000,
            revenue=5_200_000,
            ebitda=880_000,
            description="Licensed fire sprinkler installation and mandatory annual inspection services. State licensing creates strong barrier. 20 technicians with training program in place.",
            url="https://www.bizbuysell.com/business-opportunity/example/1235",
            year_established=2001,
            employees=24,
        ),
        Deal(
            title="Pacific Fresh-Cut Produce",
            location="Salinas, CA",
            asking_price=4_500_000,
            revenue=11_000_000,
            ebitda=1_200_000,
            description="USDA-inspected fresh-cut vegetable processing for grocery chains. Cold chain infrastructure, automated wash lines. 35 production workers, minimal experience required. Long-term contracts.",
            url="https://www.bizbuysell.com/business-opportunity/example/1236",
            year_established=2005,
            employees=42,
        ),
        Deal(
            title="Heritage Hide & Leather",
            location="Gloversville, NY",
            asking_price=1_800_000,
            revenue=4_100_000,
            ebitda=520_000,
            description="One of the last remaining domestic hide tanning operations. EPA-permitted facility. Processes raw hides for leather goods manufacturers. 18 floor workers trained on-site.",
            url="https://www.bizbuysell.com/business-opportunity/example/1237",
            year_established=1962,
            employees=22,
        ),
    ]

    # Process deals
    for deal in sample_deals:
        process_deal(deal)

    # Generate outputs
    excel_file = write_deals_to_excel(sample_deals)
    print(f"Excel tracker: {excel_file}")

    digest_html = generate_email_digest(sample_deals)
    with open("weekly_digest.html", "w") as f:
        f.write(digest_html)
    print("Weekly digest: weekly_digest.html")

    intro_html = generate_intro_email()
    with open("intro_email.html", "w") as f:
        f.write(intro_html)
    print("Intro email: intro_email.html")
